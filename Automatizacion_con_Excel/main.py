import os
from openpyxl import load_workbook

base_dir = os.path.dirname(__file__)
ruta_archivo = os.path.join(base_dir, "ventas.xlsx")

# Cargar archivo
archivo = load_workbook(ruta_archivo)

# Seleccionar hoja
hoja = archivo.active

# Leer datos
for fila in hoja.iter_rows(values_only=True):
    print(fila)